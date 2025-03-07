import csv
import os
import re
from typing import TextIO
import xml.etree.ElementTree as ET

import pandas as pd
from io import TextIOWrapper
from glob import glob
from itertools import dropwhile
from tqdm import tqdm
from datetime import datetime as dt, datetime
from contextlib import redirect_stderr
from importer.helper import message

from chardet.universaldetector import UniversalDetector
from core.mapping import Mapping, MappingList
from core.models import (
    BarcodeSpecification,
    Measurement,
    MeasurementFeature,
    Plate,
    PlateDimension,
    PlateMapping,
    Well,
    MappingError,
    MeasurementAssignment,
    Experiment,
    PlateDetail,
    WellDetail,
    ExperimentDetail,
    WellType,
)
from core.config import Config
from django.core.files import File
from django.utils import timezone as tz
from .helper import row_col_from_name
from helpers.logger import logger
from openpyxl import load_workbook


from datetime import timezone

GLOBAL_NOW = dt.now(timezone.utc)


def convert_string_to_datetime(date_str, time_str):
    try:
        formatted_date_str = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
        formatted_time_str = f"{time_str[:2]}:{time_str[2:4]}:{time_str[4:]}"
        combined_str = f"{formatted_date_str} {formatted_time_str}"
        datetime_obj = dt.strptime(combined_str, "%Y-%m-%d %H:%M:%S")

        datetime_obj = datetime_obj.replace(tzinfo=timezone.utc)

        return datetime_obj.isoformat()
    except ValueError as e:

        logger.warning(
            f"Cannot convert {date_str} {time_str} to datetime: {e}. The current time will be used instead."
        )
        return GLOBAL_NOW.isoformat()


def convert_sci_to_float(sci_str):
    try:
        return float(sci_str)
    except ValueError:
        logger.error(f"Cannot convert {sci_str} to float")
        return None


class BaseMapper:
    @staticmethod
    def get_files(_glob: str) -> list[str]:
        """Get all files in the given path that match the glob pattern"""
        return glob(_glob, recursive=True)

    def run(self, _glob, **kwargs):
        """Run the mapper"""

        for filename in self.get_files(_glob):
            message(
                f"Processing file {filename}...", "info", kwargs.get("room_name", None)
            )
            with redirect_stderr(None):
                detector = UniversalDetector()
                with open(filename, "rb") as file:
                    for line in file:
                        detector.feed(line)
                        if detector.done:
                            break
                    detector.close()
                encoding = detector.result.get("encoding")
            if filename.endswith(".xlsx") or filename.endswith(".txt"):
                logger.info(f"Processing {filename} as Excel file")
                data = self.parse(filename, **kwargs)
            else:
                xml_file = filename.endswith(".xml")
                # add xml_file to **kwargs
                kwargs.update({"xml_file": xml_file})
                with open(filename, "r", encoding=encoding) as file:
                    ret = self.parse(file, **kwargs)
                    if isinstance(ret, tuple):
                        data = ret[0]
                        kwargs.update(ret[1])
                    else:
                        data = ret
            kwargs.update({"filename": filename})
            self.map(data, **kwargs)

        message(
            "Refreshing materialized views...", "info", kwargs.get("room_name", None)
        )
        PlateDetail.refresh(concurrently=True)
        WellDetail.refresh(concurrently=True)
        ExperimentDetail.refresh(concurrently=True)

    def parse(self, file: TextIOWrapper | str | TextIO, **kwargs):
        raise NotImplementedError

    def map(self, data: list[dict], **kwargs) -> None:
        raise NotImplementedError

    def create_measurement_assignment(self, plate, filename):
        with open(filename, "rb") as file:
            assignment, _ = MeasurementAssignment.objects.update_or_create(
                status="success",
                plate=plate,
                filename=filename,
                measurement_file=File(file, os.path.basename(file.name)),
            )
            return assignment

    def create_plate_by_name_and_barcode(
        self,
        plate_name: str,
        plate_type: str,
        barcode: str,
        source_plate_name: str,
        **kwargs,
    ):
        try:
            barcode_prefix = barcode.split("_")[0]
            barcode_specification = BarcodeSpecification.objects.get(
                prefix=barcode_prefix
            )
        except BarcodeSpecification.DoesNotExist:
            if kwargs.get("experiment_name"):
                message(
                    f"No barcode specification found for {barcode}. Creating it.",
                    "warning",
                    kwargs.get("room_name", None),
                )

                barcode_specification, _ = BarcodeSpecification.objects.get_or_create(
                    prefix=barcode.split("_")[0],
                    sides=["North"],
                    number_of_plates=4,
                    experiment=Experiment.objects.get(
                        name=kwargs.get("experiment_name")
                    ),
                )
            else:
                message(
                    f"No barcode specification found for {barcode} and no experiment name is provided. Please provide the experiment name in order to create the missing barcode specifications.",
                    "error",
                    kwargs.get("room_name", None),
                )
                raise ValueError(
                    f"No barcode specification found for {barcode} and no experiment name is "
                    f"provided."
                    f" Please provide the experiment name in order to create the missing barcode "
                    f"specifications."
                )

        return Plate.objects.create(
            barcode=barcode,
            experiment=barcode_specification.experiment,
            dimension=self.get_plate_dimension(
                plate_name, plate_type, source_plate_name, kwargs.get("room_name")
            ),
        )

    def get_plate_dimension(
        self, plate_name: str, plate_type: str, source_plate_name, room_name
    ):
        try:
            rows, cols = row_col_from_name(
                f"{plate_type} {plate_name} {source_plate_name}"
            )
            plate_dimension = PlateDimension.objects.get(rows=rows, cols=cols)
            return plate_dimension
        except ValueError:
            message(
                f"Could not determine plate dimensions for {plate_name}",
                "error",
                room_name,
            )
            raise
        except PlateDimension.DoesNotExist:
            message(f"No plate dimension found", "error", room_name)
            raise ValueError(f"No plate dimension found for {plate_name}")


class EchoMapper(BaseMapper):
    DEFAULT_COLUMNS = Config.current.importer.echo.default.columns

    def __fast_forward_to_header_row(self, file, headers):
        """
        Fast forward to header column determined
        by finding the first column name
        """
        pattern = list(headers.values())[0]
        return dropwhile(lambda line: pattern not in line, file)

    def parse(self, file: TextIOWrapper, **kwargs) -> list[dict]:
        if kwargs.get("xml_file"):
            tree = ET.parse(file)
            root = tree.getroot()
            result = []
            source_plate_name = None
            source_plate_barcode = None
            destination_plate_name = None
            destination_plate_barcode = None
            for plate in root.find("plateInfo"):
                if plate.get("type") == "source":
                    source_plate_name = plate.get("name")
                    source_plate_barcode = plate.get("barcode")
                elif plate.get("type") == "destination":
                    destination_plate_name = plate.get("name")
                    destination_plate_barcode = plate.get("barcode")
            for w in root.find("printmap"):
                result.append(
                    {
                        "source_plate_name": source_plate_name,
                        "source_plate_barcode": source_plate_barcode,
                        "destination_plate_name": destination_plate_name,
                        "destination_plate_barcode": destination_plate_barcode,
                        "source_well": w.get("n"),
                        "destination_well": w.get("dn"),
                        "actual_volume": w.get("vl"),
                        "current_fluid_volume": w.get("cvl"),
                        "DMSO": w.get("fc"),
                        "transfer_status": "",
                    }
                )
            print(result)
            return result

        headers = kwargs.get("headers", EchoMapper.DEFAULT_COLUMNS)
        file = self.__fast_forward_to_header_row(file, headers)
        results = []
        reader = csv.DictReader(file, delimiter=",")

        for row in reader:
            # If there are None values in any of the following keys
            # or if there is a second header column we continue
            must_keys = (
                "source_plate_barcode",
                "source_plate_type",
                "source_well",
                "destination_plate_name",
                "destination_plate_barcode",
                "destination_well",
                "actual_volume",
            )
            if any(
                [
                    row[headers.get(key)] is None
                    or row[headers.get(key)] == headers.get(key)
                    for key in must_keys
                ]
            ):
                continue

            res_dict = {}
            for new_key, old_key in headers.items():
                if old_key in row:
                    res_dict[new_key] = row[old_key]
            results.append(res_dict)
        return results

    def map(self, data: list[dict], **kwargs) -> None:
        def __debug(msg):
            if kwargs.get("debug"):
                message(msg, "debug", kwargs.get("room_name", None))

        # Plates cache by barcode
        plates = {}
        # MappingList cache by source plate barcode
        mapping_lists = {}
        # Process later queue
        queue = []

        with tqdm(
            desc="Processing mappings",
            unit="mappings",
            total=len(data),
        ) as mbar:
            for entry in data:
                source_plate_name = entry["source_plate_name"]
                source_plate_barcode = entry["source_plate_barcode"]
                destination_plate_name = entry["destination_plate_name"]
                current_fluid_volume = (
                    float(entry["current_fluid_volume"])
                    if entry["current_fluid_volume"]
                    else None
                )
                current_dmso = (
                    float(entry["DMSO"].replace("%", "")) if entry["DMSO"] else None
                )
                if "destination_plate_type" in entry:
                    destination_plate_type = entry["destination_plate_type"]
                else:
                    destination_plate_type = ""
                destination_plate_barcode = entry["destination_plate_barcode"]

                if source_plate_barcode in plates:
                    source_plate = plates.get(source_plate_barcode)
                else:
                    try:
                        source_plate = Plate.objects.get(barcode=source_plate_barcode)
                        plates[source_plate_barcode] = source_plate
                    except Plate.DoesNotExist:
                        message(
                            f"""Source plate with barcode {entry['source_plate_barcode']} does not exist.
                            I try again later...""",
                            "warning",
                            kwargs.get("room_name", None),
                        )

                        queue.append(entry)
                        continue

                if destination_plate_barcode in plates:
                    destination_plate = plates.get(destination_plate_barcode)
                else:
                    try:
                        destination_plate = Plate.objects.get(
                            barcode=destination_plate_barcode
                        )
                    except Plate.DoesNotExist:
                        message(
                            f"Creating destination plate {destination_plate_name}, {destination_plate_type}"
                        )
                        destination_plate = self.create_plate_by_name_and_barcode(
                            destination_plate_name,
                            destination_plate_type,
                            destination_plate_barcode,
                            source_plate_name,
                            **kwargs,
                        )
                mapping_list_index = (
                    f"{source_plate_barcode}__**__{destination_plate_barcode}"
                )
                if mapping_list_index in mapping_lists:
                    mapping_list = mapping_lists.get(mapping_list_index)
                else:
                    mapping_list = MappingList(target=destination_plate)
                    mapping_lists[mapping_list_index] = mapping_list

                source_well = entry["source_well"]
                destination_well = entry["destination_well"]
                from_pos = source_plate.dimension.position(source_well)
                to_pos = destination_plate.dimension.position(destination_well)

                mapping = Mapping(
                    from_pos=from_pos,
                    to_pos=to_pos,
                    amount=float(entry["actual_volume"]),
                    status=entry["transfer_status"],
                    map_type=source_plate.is_control_plate,  # if it is true, we need to map the type
                    current_amount=current_fluid_volume,
                    current_dmso=current_dmso,
                )
                mapping_list.add(mapping)
                mbar.update(1)

        for mapping_list_index, mapping_list in mapping_lists.items():
            source_plate_barcode = mapping_list_index.split("__**__")[0]
            source_plate = plates.get(source_plate_barcode)
            message(
                f"Mapping {source_plate.barcode} -> {mapping_list.target.barcode}",
                "info",
                kwargs.get("room_name", None),
            )

            if source_plate.map(mapping_list, mapping_list.target):
                with open(kwargs["filename"], "rb") as file:
                    PlateMapping.objects.create(
                        source_plate=source_plate,
                        target_plate=mapping_list.target,
                        mapping_file=File(file, os.path.basename(file.name)),
                    )
                message(
                    f"Mapped {source_plate.barcode} -> {mapping_list.target.barcode}",
                    "info",
                    kwargs.get("room_name", None),
                )
                PlateDetail.refresh(concurrently=True)
                WellDetail.refresh(concurrently=True)
            else:
                message(
                    f"Error mapping {source_plate.barcode} -> {mapping_list.target.barcode}",
                    "error",
                    kwargs.get("room_name", None),
                )

        # If there are entries queued because the source plate did not yet exist
        # we try map those again but only once.
        MAX_TRY_QUEUE = 3
        try_queue = kwargs.get("try_queue", 0)
        if try_queue < MAX_TRY_QUEUE and len(queue) > 0:
            kwargs.update({"try_queue": try_queue + 1})
            self.map(queue, **kwargs)


class M1000Mapper(BaseMapper):
    RE_FILENAME = r"(?:(?P<date>[0-9]+)-(?P<time>[0-9]+)_)?(?P<barcode>[^\.]+)\.asc"

    RE_POS = r"^[A-Z]+[0-9]+$"
    RE_ID = r"^[^_]+_[^_]+$"
    RE_NUM = r"^[0-9\.]+$"
    RE_TAB = r"[\t\s]+"
    RE_SCIENTIFIC = r"[-+]?[0-9]*\.?[0-9]+([eE][-+]?[0-9]+)?"

    RE_DATE_OF_MEASUREMENT = r"^Date of measurement: (?P<date>[^\/]+)\/Time of measurement: (?P<time>[0-9:]+)$"
    RE_PLATE_DESCRIPTION = r"^Plate Description: (?P<description>.+)$"
    RE_META_DATA = r"^    (?P<key>[^:]+): (?P<value>.+)$"

    def determine_indexes(self, file: TextIOWrapper):
        """
        Determines the indexes for position and identifier.
        The rest of the columns are values.

        The strategy is to go through the lines of the file
        and stop as soon as we find an unambiguous line.
        """
        for line in file:
            parts = re.split(self.RE_TAB, line.strip())
            unambiguity = []
            pos_index = None
            id_index = None
            for idx, part in enumerate(parts):
                if match_pos := re.match(self.RE_POS, part):
                    pos_index = idx
                if match_id := re.match(self.RE_ID, part):
                    id_index = idx
                # if this is an unambiguous value
                unambiguity.append([match_pos, match_id])
            # if the line is an unambiguous line
            df = pd.DataFrame(unambiguity)
            if df.count().apply(lambda x: x == 1).all():
                file.seek(0)
                return pos_index, id_index
        file.seek(0)
        raise MappingError(f"File has not the desired format: {file.name}")

    def parse(self, file: TextIOWrapper, **kwargs) -> list[dict]:
        def __debug(msg):
            if kwargs.get("debug"):
                message(msg, "debug", kwargs.get("room_name", None))

        match = re.match(self.RE_FILENAME, os.path.basename(file.name))
        if match:
            barcode = match.group("barcode")
        else:
            raise MappingError(
                f"File name {os.path.basename(file.name)} does not match conventions."
            )

        results = []
        measurement_date = tz.now()
        plate_description = None
        meta_data = [{}]
        meta_data_idx = 0
        pos_index, id_index = self.determine_indexes(file)
        for line in file:
            parts = re.split(self.RE_TAB, line)
            # For a value line we expect at least 3 parts (pos, id, value)
            # and the position part should match the position regex
            if len(parts) >= 3 and re.match(self.RE_POS, parts[pos_index]):
                position = parts[pos_index]
                identifier = parts[id_index]
                result = {
                    "position": position,
                    "identifier": identifier,
                    "values": [
                        convert_sci_to_float(part)
                        for idx, part in enumerate(parts)
                        if idx not in [pos_index, id_index]
                        and (
                            re.match(self.RE_NUM, part)
                            or re.match(self.RE_SCIENTIFIC, part)
                        )
                    ],
                }
                # If there are no values in this line we ignore it
                if len(result["values"]) > 0:
                    __debug(f"result: {result}")
                    results.append(result)
            elif match := re.match(self.RE_DATE_OF_MEASUREMENT, line):
                measurement_date = dt.strptime(
                    f"{match.group('date')} {match.group('time')}", "%Y-%m-%d %H:%M:%S"
                )
            elif match := re.match(self.RE_PLATE_DESCRIPTION, line):
                plate_description = match.group("description")
            elif match := re.match(self.RE_META_DATA, line):
                if match.group("key") in meta_data[meta_data_idx]:
                    meta_data_idx += 1
                    meta_data.append({})
                meta_data[meta_data_idx][match.group("key")] = match.group("value")
            else:
                pass

        kwargs.update(
            {
                "barcode": barcode,
                "measurement_date": measurement_date,
                "plate_description": plate_description,
                "meta_data": list(reversed(meta_data)),  # the order of the
                # metadata in the file is reversed compared to the order of the values
            }
        )
        return results, kwargs

    def map(self, data: list[dict], **kwargs) -> None:
        def __debug(msg):
            if kwargs.get("debug"):
                message(msg, "debug", kwargs.get("room_name", None))

        barcode = kwargs.get("barcode")
        try:
            plate = Plate.objects.get(barcode=barcode)
        except Plate.DoesNotExist:

            message(
                f"Plate with barcode {barcode} does not exist. Creating it.",
                "warning",
                kwargs.get("room_name", None),
            )

            barcode_specification, _ = BarcodeSpecification.objects.get_or_create(
                prefix=barcode.split("_")[0],
                sides=["North"],
                number_of_plates=4,
                experiment=Experiment.objects.get(name=kwargs.get("experiment_name")),
            )
            plate = Plate.objects.create(
                barcode=barcode,
                dimension=PlateDimension.by_num_wells(len(data)),
                experiment=barcode_specification.experiment,
            )

        with tqdm(
            desc="Processing measurements",
            unit="measurement",
            total=len(data),
        ) as mbar:
            assignment = self.create_measurement_assignment(
                plate, kwargs.get("filename")
            )

            for entry in data:
                __debug(f"Entry: {entry}")
                position = plate.dimension.position(entry.get("position"))
                well = plate.well_at(position)
                if not well:
                    well = Well.objects.create(plate=plate, position=position)

                measurement_names = None
                if kwargs.get("measurement_name"):
                    measurement_names = kwargs.get("measurement_name").split(",")
                for idx, value in enumerate(entry.get("values")):
                    if measurement_names:
                        label = measurement_names[idx]
                    else:
                        label = kwargs.get("meta_data")[idx].get("Label")

                    Measurement.objects.update_or_create(
                        well=well,
                        label=label,
                        measured_at=kwargs.get("measurement_date"),
                        defaults={
                            "value": value,
                            "identifier": entry.get("identifier"),
                            "measurement_assignment": assignment,
                        },
                    )
                mbar.update(1)


class MicroscopeMapper(BaseMapper):
    RE_FILENAME = (
        r"(?P<date>\d+)[-_](?P<time>\d+)[-_](?P<barcode>[^\.]+)\.(?P<ext>xlsx|txt)$"
    )

    def parse(self, file, **kwargs):
        filename = file
        basename = os.path.basename(filename)
        match = re.match(self.RE_FILENAME, basename)
        if match:
            date = match.group("date")
            time = match.group("time")
            barcode = match.group("barcode")
            ext = match.group("ext")
        else:
            message(
                f"Filename {basename} does not match expected pattern.",
                "error",
                kwargs.get("room_name", None),
            )
            raise ValueError(f"Filename {basename} does not match expected pattern.")

        kwargs.update(
            {"filename": filename, "barcode": barcode, "date": date, "time": time}
        )

        if ext == "xlsx":
            wb = load_workbook(file)
            sheet = wb.active
            metadata = self.__parse_metadata(sheet)
            results = self.__parse_results(sheet)
            layout = self.__parse_layout(sheet, len(results))
        elif ext == "txt":
            content = open(file, "r")
            lines = [line.strip() for line in content.readlines() if line.strip()]
            metadata = self.__parse_metadata_txt(lines)
            results = self.__parse_results_txt(lines)
            layout = (
                {}
            )  # Implement __parse_layout_txt if layout info is present in .txt files
            # Prefer date and time from metadata if available
            date = metadata.get("Date", date)
            time = metadata.get("Time", time)
            logger.info(f"Date: {date}, Time: {time}")
            content.close()
        else:
            raise ValueError(f"Unsupported file extension: {ext}")

        return {
            "metadata": metadata,
            "results": results,
            "date": date,
            "time": time,
            "barcode": barcode,
            "layout": layout,
        }

    def map(self, data: dict, **kwargs) -> None:

        RE_NUMBER = r"^[0-9]+(\.[0-9]+)?$"
        RE_SCIENCE = r"^[0-9\.]+[eE][+-]?[0-9]+$"

        barcode = data["barcode"]
        try:
            plate = Plate.objects.get(barcode=barcode)
        except Plate.DoesNotExist:
            message(
                f"Plate with barcode {barcode} does not exist. Creating it.",
                "warning",
                kwargs.get("room_name", None),
            )
            barcode_specification, _ = BarcodeSpecification.objects.get_or_create(
                prefix=barcode.split("_")[0],
                sides=["North"],
                number_of_plates=4,
                experiment=Experiment.objects.get(name=kwargs.get("experiment_name")),
            )
            plate = Plate.objects.create(
                barcode=barcode,
                dimension=PlateDimension.by_num_wells(len(data["results"])),
                experiment=barcode_specification.experiment,
            )
        measured_at = convert_string_to_datetime(data["date"], data["time"])
        with tqdm(
            desc="Processing microscope output",
            unit="measurement",
            total=len(data["results"]),
        ) as mbar:
            for entry in data["results"]:
                if not entry.get("Well") or entry.get("Well") == "Well":
                    continue

                position = plate.dimension.position(entry.get("Well"))

                well = plate.well_at(position)
                if not well:

                    well = Well.objects.create(plate=plate, position=position)
                if data["layout"] and entry.get("Well") in data["layout"]:
                    well_type = data["layout"][entry.get("Well")]
                    well.type = WellType.objects.get(name=well_type)
                    well.save()
                for key, value in entry.items():
                    if key in ["Well ID", "Well"]:
                        continue
                    if re.match(RE_NUMBER, str(value)):
                        value = float(value)
                    elif re.match(RE_SCIENCE, str(value)):
                        value = convert_sci_to_float(value)
                    else:
                        continue
                    Measurement.objects.update_or_create(
                        well=well,
                        label=key,
                        measured_at=measured_at,
                        defaults={
                            "value": value,
                        },
                    )

                mbar.update(1)
        self.create_measurement_assignment(plate, kwargs.get("filename"))

    def __parse_metadata(self, sheet):
        metadata = {}
        current_label = None
        for row in sheet.iter_rows(min_row=1, max_row=38):
            for index, cell in enumerate(row):
                if index == 0 and cell.value:
                    current_label = str(cell.value).strip().lstrip().replace(":", "")
                    metadata[current_label] = []

                elif index != 0 and cell.value:
                    str_value = str(cell.value)
                    if ": " in str_value:
                        k = str_value.split(":")[0].strip().lstrip()
                        v = str_value.split(":")[1].strip().lstrip()
                        metadata[k] = v
                    else:
                        metadata[current_label].append(str_value)
                if str(cell.value).strip().lstrip().lower() == "results":
                    break

        return metadata

    def __parse_results(self, sheet):
        results_data = []
        headers = []
        results_start_row = None
        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if (
                row[1]
                and str(row[1]).lower() == "well id"
                or str(row[1]).lower() == "well"
            ):
                for cell in sheet.iter_rows(
                    min_row=index + 1,
                    max_row=index + 1,
                    values_only=True,  # we use index +1 because indices start with 1 in Excel
                ):
                    if cell:
                        headers.append(cell)
                results_start_row = index + 2
                break
        headers = [header for header in headers[0] if header is not None]
        if headers and results_start_row:
            for row in sheet.iter_rows(min_row=results_start_row, values_only=True):
                if not any(row):
                    continue
                row_data = dict(zip(headers, row[1:]))
                results_data.append(row_data)

        return results_data

    def __parse_layout(self, sheet, plate):
        layout_start_row = None
        layout_end_row = None
        layout_data = []
        position_type = {}
        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if row[0] and str(row[0]).lower() == "layout":
                layout_start_row = index + 2
            if row[0] and str(row[0]).lower() == "results":
                layout_end_row = index - 1
        if layout_start_row and layout_end_row:
            for row in sheet.iter_rows(
                min_row=layout_start_row, max_row=layout_end_row, values_only=True
            ):
                if not any(row):
                    continue
                layout_data.append(row[1:])

        for item in layout_data:
            if item[0]:
                for index, value in enumerate(item):
                    if value in ["POS", "NEG"]:
                        well_position = f"{item[0]}{index}"
                        well_type = "P" if value == "POS" else "N"
                        position_type[well_position] = well_type

        return position_type

    def __parse_metadata_txt(self, lines):
        metadata = {}
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            if line == "Results":
                break  # Stop parsing metadata when 'Results' is reached
            if not line:
                i += 1
                continue
            if "\t" in line:
                parts = line.split("\t", 1)
                if len(parts) == 2:
                    key, value = parts
                    metadata[key.strip()] = value.strip()
            elif ":" in line:
                parts = line.split(":", 1)
                if len(parts) == 2:
                    key, value = parts
                    metadata[key.strip()] = value.strip()

            i += 1

        return metadata

    def __parse_results_txt(self, lines):
        results = []
        # Find the index where 'Results' section starts
        i = 0
        while i < len(lines):
            if lines[i] == "Results":
                i += 1  # Skip the 'Results' header
                break
            i += 1

        # Now parse the results
        while i < len(lines):
            line = lines[i]
            if line.strip() == "":
                break
            parts = line.split("\t")
            if len(parts) == 2:
                well, lum = parts
                if well.strip() == "Well":
                    i += 1
                    continue
                results.append({"Well": well.strip(), "Lum": lum.strip()})
            i += 1

        print(results[0])
        return results


class DatMapper(BaseMapper):
    def parse(self, file: TextIOWrapper | str | TextIO, **kwargs):

        message(
            "Parsing DAT file... ------------------------------------------", "debug"
        )

        content = file.read()
        pattern = r"Date:\s*(\d{2}/\d{2}/\d{4})\s*Time:\s*(\d{2}:\d{2}:\d{2})"
        match = re.search(pattern, content)

        if match:
            date_str = match.group(1)
            time_str = match.group(2)
            datetime_str = f"{date_str} {time_str}"
            datetime_obj = datetime.strptime(datetime_str, "%d/%m/%Y %H:%M:%S")
        else:
            datetime_obj = datetime(
                2024, 4, 4, 11, 00
            )  # random time as a quick and dirty solution for now
        last_counts = content.split("Chromatic / Channel:")[-1].split("\n")[3:]
        result = {"measured_at": datetime_obj, "counts": []}
        for line in last_counts:
            line_list = line.split("\t")
            if len(line_list) > 0:
                for i in line_list:
                    if i:
                        result["counts"].append(int(i))
        return result

    def map(self, data: dict, **kwargs) -> None:
        filename = kwargs.get("filename")
        barcode = filename.split("/")[-1].split(".")[0]
        try:
            plate = Plate.objects.get(barcode=barcode)
        except Plate.DoesNotExist:
            message(
                f"Plate with barcode {barcode} does not exist. Creating it.",
                "warning",
                kwargs.get("room_name", None),
            )
            barcode_specification, _ = BarcodeSpecification.objects.get_or_create(
                prefix=barcode.split("_")[0],
                sides=["North"],
                number_of_plates=4,
                experiment=Experiment.objects.get(name=kwargs.get("experiment_name")),
            )
            plate = Plate.objects.create(
                barcode=barcode,
                dimension=PlateDimension.by_num_wells(len(data["counts"])),
                experiment=barcode_specification.experiment,
            )
        with tqdm(
            desc="Processing dat file output",
            unit="measurement",
            total=len(data),
        ) as mbar:
            for idx, value in enumerate(data["counts"]):
                well = plate.well_at(idx)
                if not well:
                    well = Well.objects.create(plate=plate, position=idx)
                Measurement.objects.update_or_create(
                    well=well,
                    label="value",
                    measured_at=data["measured_at"],
                    defaults={
                        "value": value,
                    },
                )
                mbar.update(1)
        self.create_measurement_assignment(plate, kwargs.get("filename"))
